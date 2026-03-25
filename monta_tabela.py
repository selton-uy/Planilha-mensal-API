import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from busca_dados import consultar_processo, extrair_dados
import os

COLUNA_CODIGO    = 'Processo (padrão sistema vivo)'
COLUNA_RECORRENTE = 'Recorrente (Vivo, Parte Autora, Ambos)'
COLUNA_PARTE_AUTORA = 'Parte autora'
FILL_VERMELHO = PatternFill(start_color="FF4C4C", end_color="FF4C4C", fill_type="solid")


def _pintar_falhas(caminho_saida: str, linhas_falhas: list, colunas_header: list):
    """Abre o xlsx salvo e pinta de vermelho as células de 'Parte autora' que falharam."""
    if not linhas_falhas:
        return

    wb = load_workbook(caminho_saida)
    ws = wb.active

    # Descobre o índice da coluna no Excel (1-based)
    try:
        col_idx = colunas_header.index(COLUNA_PARTE_AUTORA) + 1
    except ValueError:
        return  # coluna não existe, não faz nada

    for linha_excel in linhas_falhas:
        ws.cell(row=linha_excel, column=col_idx).fill = FILL_VERMELHO

    wb.save(caminho_saida)


def pecorrer_tabela(tabela: pd.DataFrame, coluna_codigo: str, caminho_saida: str, log):
    total = len(tabela)
    falhas = []  # guarda as linhas do Excel que ficaram em branco

    try:
        for index, row in tabela.iterrows():
            try:
                codigo = str(row[coluna_codigo])
                if not codigo or codigo.lower() == 'nan':
                    continue

                # Pula se recorrente já estiver preenchido
                recorrente_atual = str(row.get(COLUNA_PARTE_AUTORA, '')).strip()
                if recorrente_atual and recorrente_atual.lower() != 'nan':
                    log(f"⏭️  [{str(int(index)+1).zfill(2)}] Já preenchido, pulando: {codigo}")
                    continue

                processo = consultar_processo(codigo)
                dados_processo = extrair_dados(processo)

                recorrente = str(dados_processo.get('RECORRENTE') or "").upper()
                parte_autora = dados_processo.get('RECORRENTE')

                tabela.at[index, COLUNA_PARTE_AUTORA]                        = parte_autora
                tabela.at[index, 'Turma/Câmara']                              = dados_processo.get('TURMA')
                tabela.at[index, 'Relator - Magistrado - 2ª Instância']       = dados_processo.get('RELATOR')
                tabela.at[index, 'Fundamentação principal']                   = dados_processo.get('SUMULA')

                if 'TELEFONICA' in recorrente and 'E OUTRO' in recorrente:
                    resultado = 'AMBOS'
                elif 'TELEFONICA' in recorrente:
                    resultado = 'VIVO'
                elif recorrente == '':
                    resultado = ''
                else:
                    resultado = 'PARTE AUTORA'

                tabela.at[index, COLUNA_RECORRENTE] = resultado

                # Marca como falha se parte autora ficou vazia
                if not parte_autora:
                    linha_excel = int(index) + 2  # +1 header, +1 porque Excel é 1-based
                    falhas.append(linha_excel)
                    log(f"⚠️  [{str(int(index)+1).zfill(2)}/{total}] Sem dados: {codigo}")
                else:
                    log(f"✅ [{str(int(index)+1).zfill(2)}/{total}] Sucesso: {codigo}")

            except Exception as e:
                log(f"❌ [{str(int(index)+1).zfill(2)}] Erro em {row[coluna_codigo]}: {e}")
                continue

    except KeyboardInterrupt:
        log("🛑 Processo interrompido pelo usuário.")
        _salvar_backup(tabela, caminho_saida, log)
        raise
    except Exception as e:
        log(f"💥 Erro crítico: {e}")
        _salvar_backup(tabela, caminho_saida, log)
        raise

    tabela.to_excel(caminho_saida, index=False)

    # Pinta as falhas de vermelho
    _pintar_falhas(caminho_saida, falhas, list(tabela.columns))

    # Resumo final
    log(f"💾 Planilha salva em: {caminho_saida}")
    if falhas:
        log(f"⚠️  {len(falhas)} processo(s) ficaram sem dados e estão marcados em vermelho.")
    else:
        log("🎉 Todos os processos foram preenchidos com sucesso!")

    return len(falhas)


def _salvar_backup(tabela: pd.DataFrame, caminho_saida: str, log):
    backup = caminho_saida.replace('.xlsx', '_backup.xlsx')
    tabela.to_excel(backup, index=False)
    log(f"⚠️ Backup salvo em: {backup}")


def main(caminho_entrada: str, caminho_saida: str, log=print):
    log(f"📂 Lendo: {caminho_entrada}")
    tabela = pd.read_excel(caminho_entrada)
    tabela = tabela.astype(str)
    tabela = tabela.replace('nan', '')
    log(f"📋 {len(tabela)} registros encontrados. Iniciando consultas...")
    return pecorrer_tabela(tabela, COLUNA_CODIGO, caminho_saida, log)
